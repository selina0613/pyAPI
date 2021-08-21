
from openpyxl import load_workbook
from tools.handle_path import case_data_dir
from conf.setting import register_case_data

class HangleExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_test_case(self):
        wb = load_workbook(filename=self.file_name) # 实例化excel
        sh = wb[self.sheet_name] # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = all_excel_data[0]  # 获取表头
        case_data_list = all_excel_data[1:] # 获取所有的用例数据
        test_case_list = [] # 新建空列表，收集所有的测试用例数据
        for case in case_data_list:
            #print('表头',excel_title)
            #print('测试用例',case)
            test_case = dict(zip(excel_title,case))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            test_case_list.append(test_case) # 将每次拼接好的测试用例添加到test_case_list
            #print("拼接后的数据：",test_case)
        #print(len(test_case_list))
        wb.close()
        return test_case_list

if __name__ == '__main__':
    cl = HangleExcel(case_data_dir, register_case_data["sheet_name"])
    result = cl.get_test_case()
    print(result)